#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Batch ToxinPred3 toxicity prediction for Excel peptides.
Predicts all peptides, then writes results to both Excel files.
"""
import subprocess
import re
import os
import time
import sys
import tempfile
from collections import OrderedDict

import openpyxl

# ==================== CONFIG ====================
TOXINPRED_URL = "https://webs.iiitd.edu.in/raghava/toxinpred3/prediction_action.php"
DISP_URL = "https://webs.iiitd.edu.in/raghava/toxinpred3/disp1.php"
THRESHOLD = "0.38"
MODEL = "1"  # Hybrid
BATCH_SIZE = 500  # Website limit per submission
MAX_WAIT = 300  # Max seconds to wait per batch
POLL_INTERVAL = 5

FILE1 = r"D:\kanmao\桌面\活性大于等于0.4.xlsx"
FILE2 = r"D:\kanmao\桌面\digestion_results_normalized.xlsx"
SHEET_NAME = "digestion_results"

# ==================== HELPERS ====================

def make_fasta(peptides):
    lines = []
    for i, seq in enumerate(peptides, 1):
        lines.append(">Seq_%d" % i)
        lines.append(seq)
    return "\n".join(lines)


def submit_and_fetch(fasta_text):
    """Submit batch via curl, poll until results ready, return parsed list"""
    tmpfile = os.path.join(tempfile.gettempdir(), "toxinpred_batch.fasta")
    with open(tmpfile, "w", encoding="utf-8") as f:
        f.write(fasta_text)

    # Submit
    cmd = [
        "curl", "-s", "-X", "POST", TOXINPRED_URL,
        "--form", "seq=<%s" % tmpfile,
        "--form", "terminus=%s" % MODEL,
        "--form", "th=%s" % THRESHOLD,
        "--form", "submit=Submit",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    match = re.search(r"ran=(\d+)", result.stdout)
    if not match:
        print("  ERROR: Could not extract Job ID")
        return None
    ran = match.group(1)

    # Poll for results
    elapsed = 0
    while elapsed < MAX_WAIT:
        time.sleep(POLL_INTERVAL)
        elapsed += POLL_INTERVAL
        cmd2 = ["curl", "-s", "%s?ran=%s" % (DISP_URL, ran)]
        res2 = subprocess.run(cmd2, capture_output=True, text=True, timeout=60)
        if "<td>" in res2.stdout:
            # Parse
            tbody = re.search(r"<tbody>(.*?)</tbody>", res2.stdout, re.DOTALL)
            if not tbody:
                continue
            rows = re.findall(r"<tr>(.*?)</tr>", tbody.group(1), re.DOTALL)
            results = {}
            for row in rows:
                cells = re.findall(r"<td>(.*?)</td>", row, re.DOTALL)
                if len(cells) >= 7:
                    seq = cells[1].strip()
                    results[seq] = {
                        "ml_score": float(cells[2].strip()),
                        "hybrid_score": float(cells[5].strip()),
                        "prediction": cells[6].strip(),
                        "ppv": float(cells[7].strip()) if len(cells) >= 8 else 0,
                    }
            return results

    print("  TIMEOUT after %ds" % MAX_WAIT)
    return None


# ==================== MAIN ====================

def main():
    print("=" * 60)
    print("ToxinPred3 Batch Toxicity Prediction")
    print("=" * 60)

    # ---- Step 1: Read both files, get unique peptides ----
    print("\n[1] Reading Excel files...")
    wb1 = openpyxl.load_workbook(FILE1)
    ws1 = wb1[SHEET_NAME]
    wb2 = openpyxl.load_workbook(FILE2)
    ws2 = wb2[SHEET_NAME]

    # Collect all unique peptides to predict
    all_peptides = set()
    for ws in [ws1, ws2]:
        for row in ws.iter_rows(min_row=2, min_col=10, max_col=10, values_only=True):
            seq = row[0]
            if seq:
                all_peptides.add(str(seq).strip().upper())

    all_peptides = sorted(all_peptides)
    print("  Unique peptides to predict: %d" % len(all_peptides))

    # Check if we have cached results
    cache_file = os.path.join(tempfile.gettempdir(), "toxinpred_cache.txt")
    already_done = {}
    if os.path.exists(cache_file):
        with open(cache_file) as f:
            for line in f:
                parts = line.strip().split("\t")
                if len(parts) >= 5:
                    already_done[parts[0]] = {
                        "ml_score": float(parts[1]),
                        "hybrid_score": float(parts[2]),
                        "prediction": parts[3],
                        "ppv": float(parts[4]),
                    }
        print("  Already predicted (cached): %d" % len(already_done))

    pending = [p for p in all_peptides if p not in already_done]
    print("  Remaining to predict: %d" % len(pending))

    # ---- Step 2: Batch submit ----
    if pending:
        batches = [pending[i:i+BATCH_SIZE] for i in range(0, len(pending), BATCH_SIZE)]
        print("\n[2] Submitting %d batches (up to %d peptides each)..." % (len(batches), BATCH_SIZE))

        for bi, batch in enumerate(batches, 1):
            print("\n--- Batch %d/%d (%d peptides) ---" % (bi, len(batches), len(batch)))
            fasta = make_fasta(batch)
            results = submit_and_fetch(fasta)
            if results:
                already_done.update(results)
                print("  Got %d results" % len(results))
                # Save cache incrementally
                with open(cache_file, "w") as f:
                    for seq, r in already_done.items():
                        f.write("%s\t%.3f\t%.3f\t%s\t%.3f\n" % (
                            seq, r["ml_score"], r["hybrid_score"], r["prediction"], r["ppv"]
                        ))
            else:
                print("  FAILED for this batch")
            # Small delay between batches
            if bi < len(batches):
                time.sleep(3)

    print("\n  Total predicted: %d peptides" % len(already_done))

    # ---- Step 3: Write results back to Excel ----
    print("\n[3] Writing results to Excel files...")

    # Helper to add columns and fill values
    def add_toxicity_columns(ws):
        """Add 4 toxicity columns after existing columns and fill with data"""
        # Add headers (row 1)
        headers_added = False
        toxicity_cols = {}  # col_index -> col_name
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val and str(val).strip() in ("Toxicity_Prediction", "Toxicity_Hybrid_Score",
                                              "Toxicity_ML_Score", "Toxicity_PPV"):
                toxicity_cols[str(val).strip()] = col_idx

        if not toxicity_cols:
            # Add new columns
            next_col = ws.max_column + 1
            headers = ["Toxicity_Prediction", "Toxicity_Hybrid_Score",
                       "Toxicity_ML_Score", "Toxicity_PPV"]
            for i, h in enumerate(headers):
                ws.cell(row=1, column=next_col + i, value=h)
            toxicity_cols = {
                "Toxicity_Prediction": next_col,
                "Toxicity_Hybrid_Score": next_col + 1,
                "Toxicity_ML_Score": next_col + 2,
                "Toxicity_PPV": next_col + 3,
            }

        return toxicity_cols

    # Process file 1
    tc1 = add_toxicity_columns(ws1)
    filled1 = 0
    for row_idx in range(2, ws1.max_row + 1):
        seq = ws1.cell(row=row_idx, column=10).value
        if seq:
            seq = str(seq).strip().upper()
            if seq in already_done:
                r = already_done[seq]
                ws1.cell(row=row_idx, column=tc1["Toxicity_Prediction"], value=r["prediction"])
                ws1.cell(row=row_idx, column=tc1["Toxicity_Hybrid_Score"], value=r["hybrid_score"])
                ws1.cell(row=row_idx, column=tc1["Toxicity_ML_Score"], value=r["ml_score"])
                ws1.cell(row=row_idx, column=tc1["Toxicity_PPV"], value=r["ppv"])
                filled1 += 1
    print("  File 1: filled %d / %d rows" % (filled1, ws1.max_row - 1))
    wb1.save(FILE1)
    print("  Saved: %s" % FILE1)

    # Process file 2
    tc2 = add_toxicity_columns(ws2)
    filled2 = 0
    for row_idx in range(2, ws2.max_row + 1):
        seq = ws2.cell(row=row_idx, column=10).value
        if seq:
            seq = str(seq).strip().upper()
            if seq in already_done:
                r = already_done[seq]
                ws2.cell(row=row_idx, column=tc2["Toxicity_Prediction"], value=r["prediction"])
                ws2.cell(row=row_idx, column=tc2["Toxicity_Hybrid_Score"], value=r["hybrid_score"])
                ws2.cell(row=row_idx, column=tc2["Toxicity_ML_Score"], value=r["ml_score"])
                ws2.cell(row=row_idx, column=tc2["Toxicity_PPV"], value=r["ppv"])
                filled2 += 1
    print("  File 2: filled %d / %d rows" % (filled2, ws2.max_row - 1))
    wb2.save(FILE2)
    print("  Saved: %s" % FILE2)

    # ---- Summary ----
    toxins = sum(1 for r in already_done.values() if r["prediction"] == "Toxin")
    nontoxins = sum(1 for r in already_done.values() if r["prediction"] == "Non-Toxin")
    print("\n" + "=" * 60)
    print("DONE: %d Toxin | %d Non-Toxin (total %d peptides)" %
          (toxins, nontoxins, len(already_done)))
    print("=" * 60)


if __name__ == "__main__":
    main()
