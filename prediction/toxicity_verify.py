#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verify ToxinPred3 results in both Excel files match the cache."""
import os
import random
import openpyxl

random.seed(42)

FILES = [
    r"D:\kanmao\桌面\活性大于等于0.4.xlsx",
    r"D:\kanmao\桌面\digestion_results_normalized.xlsx",
]
CACHE = r"C:\Users\kanmao\AppData\Local\Temp\claude\toxinpred_cache.txt"
if not os.path.exists(CACHE):
    CACHE = r"/tmp/claude/toxinpred_cache.txt"

# Load cache
cache = {}
with open(CACHE) as f:
    for line in f:
        parts = line.strip().split("\t")
        if len(parts) >= 5:
            cache[parts[0]] = {
                "ml_score": float(parts[1]),
                "hybrid_score": float(parts[2]),
                "prediction": parts[3],
                "ppv": float(parts[4]),
            }

for fname in FILES:
    print("=" * 70)
    print("File: %s" % os.path.basename(fname))
    print("=" * 70)

    wb = openpyxl.load_workbook(fname)
    ws = wb["digestion_results"]

    # Find column indices
    header = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h:
            header[h] = c

    seq_col = header.get("fragment_sequence")
    pred_col = header.get("Toxicity_Prediction")
    score_col = header.get("Toxicity_Hybrid_Score")
    ml_col = header.get("Toxicity_ML_Score")
    ppv_col = header.get("Toxicity_PPV")

    print("Columns: seq=%d, pred=%d, hybrid=%d, ml=%d, ppv=%d" %
          (seq_col, pred_col, score_col, ml_col, ppv_col))

    # ============ Check 1: Coverage ============
    total = ws.max_row - 1
    missing = 0
    empty_seq = 0
    mismatches = 0

    for row in range(2, ws.max_row + 1):
        seq = ws.cell(row=row, column=seq_col).value
        pred = ws.cell(row=row, column=pred_col).value
        if not seq:
            empty_seq += 1
            continue
        if not pred:
            missing += 1
            continue
        seq = str(seq).strip().upper()
        # Exact value check
        if seq in cache:
            exp = cache[seq]
            e_pred = exp["prediction"]
            e_score = exp["hybrid_score"]
            e_ml = exp["ml_score"]
            e_ppv = exp["ppv"]
            f_pred = str(pred)
            f_score = ws.cell(row=row, column=score_col).value
            f_ml = ws.cell(row=row, column=ml_col).value
            f_ppv = ws.cell(row=row, column=ppv_col).value
            if (f_pred != e_pred or
                abs(float(f_score) - e_score) > 0.001 or
                abs(float(f_ppv) - e_ppv) > 0.001):
                mismatches += 1
        else:
            missing += 1

    print("Total: %d | Empty seq: %d | Missing: %d | Mismatched: %d" %
          (total, empty_seq, missing, mismatches))

    # ============ Check 2: Random spot checks ============
    valid_rows = [r for r in range(2, ws.max_row + 1)
                  if ws.cell(row=r, column=seq_col).value]
    sampled = random.sample(valid_rows, min(15, len(valid_rows)))
    sampled.sort()

    print("\nRandom spot checks (15 rows):")
    all_ok = True
    for row in sampled:
        seq = str(ws.cell(row=row, column=seq_col).value).strip().upper()
        pred = ws.cell(row=row, column=pred_col).value
        score = ws.cell(row=row, column=score_col).value
        ppv = ws.cell(row=row, column=ppv_col).value

        if seq in cache:
            exp = cache[seq]
            ok = (str(pred) == exp["prediction"] and
                  abs(float(score) - exp["hybrid_score"]) < 0.001 and
                  abs(float(ppv) - exp["ppv"]) < 0.001)
            status = "OK" if ok else "MISMATCH!"
            if not ok:
                all_ok = False
            print("  [%s] row %d: %-20s -> %-10s score=%.3f ppv=%.3f" %
                  (status, row, seq, pred, score, ppv))
        else:
            print("  [MISSING] row %d: %s not found in cache" % (row, seq))
            all_ok = False

    if all_ok and mismatches == 0 and missing == 0:
        print("\n  >>> ALL VERIFIED: 100%% matched <<<")
    else:
        print("\n  >>> ERRORS FOUND <<<")

    wb.close()

print("\n" + "=" * 70)
print("Cache contains %d unique peptides" % len(cache))
toxins = sum(1 for v in cache.values() if v["prediction"] == "Toxin")
nontoxins = sum(1 for v in cache.values() if v["prediction"] == "Non-Toxin")
print("Cache: %d Toxin | %d Non-Toxin" % (toxins, nontoxins))
